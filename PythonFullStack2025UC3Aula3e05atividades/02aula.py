from openpyxl import Workbook, load_workbook

NOME_ARQUIVO_EXCEL = "funcionarios.xlsx"

def adicionar_funcionario(nome, cargo, salario):
    """Adiciona um novo funcionário à planilha."""
    try:
        wb = load_workbook(NOME_ARQUIVO_EXCEL)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Cargo", "Salário"])

    ws.append([nome, cargo, salario])
    wb.save(NOME_ARQUIVO_EXCEL)
    print(f"Funcionário '{nome}' adicionado com sucesso.")

def listar_funcionarios():
    """Lista todos os funcionários da planilha."""
    if os.path.exists(NOME_ARQUIVO_EXCEL):
        wb = load_workbook(NOME_ARQUIVO_EXCEL)
        ws = wb.active
        funcionarios = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # Começa da segunda linha para pular o cabeçalho
            if row[0]:  # Garante que a linha não está vazia (pelo nome)
                funcionarios.append({"Nome": row[0], "Cargo": row[1], "Salário": row[2]})

        if funcionarios:
            print("\nLista de Funcionários:")
            for funcionario in funcionarios:
                print(f"- Nome: {funcionario['Nome']}, Cargo: {funcionario['Cargo']}, Salário: R$ {funcionario['Salário']:.2f}")
        else:
            print("\nNenhum funcionário cadastrado.")
    else:
        print(f"\nO arquivo '{NOME_ARQUIVO_EXCEL}' não existe.")

def calcular_media_salarial():
    """Calcula e exibe a média salarial dos funcionários."""
    if os.path.exists(NOME_ARQUIVO_EXCEL):
        wb = load_workbook(NOME_ARQUIVO_EXCEL)
        ws = wb.active
        total_salarios = 0
        num_funcionarios = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is not None:  # Verifica se o salário não está vazio
                try:
                    total_salarios += float(row[2])
                    num_funcionarios += 1
                except ValueError:
                    print(f"Aviso: Salário inválido encontrado para '{row[0]}'.")

        if num_funcionarios > 0:
            media_salarial = total_salarios / num_funcionarios
            print(f"\nMédia Salarial dos Funcionários: R$ {media_salarial:.2f}")
        else:
            print("\nNenhum funcionário com salário válido encontrado para calcular a média.")
    else:
        print(f"\nO arquivo '{NOME_ARQUIVO_EXCEL}' não existe.")

def main():
    # Inicialização básica (será executada apenas na primeira vez ou se o arquivo não existir)
    if not os.path.exists(NOME_ARQUIVO_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Funcionários"
        ws.append(["Nome", "Cargo", "Salário"])
        ws.append(["Maria", "Engenheira", 8500])
        ws.append(["Pedro", "Analista", 6000])
        wb.save(NOME_ARQUIVO_EXCEL)
        print(f"Arquivo '{NOME_ARQUIVO_EXCEL}' inicializado com alguns funcionários.")

    while True:
        print("\nOpções:")
        print("1. Adicionar Funcionário")
        print("2. Listar Funcionários")
        print("3. Calcular Média Salarial")
        print("4. Sair")

        opcao = input("Escolha uma opção: ")

        if opcao == '1':
            nome = input("Digite o nome do funcionário: ")
            cargo = input("Digite o cargo do funcionário: ")
            while True:
                try:
                    salario = float(input("Digite o salário do funcionário: "))
                    break
                except ValueError:
                    print("Salário inválido. Digite um número.")
            adicionar_funcionario(nome, cargo, salario)
        elif opcao == '2':
            listar_funcionarios()
        elif opcao == '3':
            calcular_media_salarial()
        elif opcao == '4':
            print("Saindo do programa.")
            break
        else:
            print("Opção inválida. Tente novamente.")

if __name__ == "__main__":
    import os  # Importando o módulo os dentro do bloco main para evitar problemas se rodado em certos ambientes
    main()